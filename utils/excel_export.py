import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelExporter:
    """Excel export utility for reports and data"""
    
    def __init__(self):
        self.output_dir = Path.home() / "Documents" / "الحسيني" / "التقارير"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_reports(self, filename, sales_report, profit_report, inventory_report, top_products):
        """Export comprehensive reports to Excel"""
        try:
            filepath = self.output_dir / filename
            
            # Create workbook
            with pd.ExcelWriter(str(filepath), engine='openpyxl') as writer:
                
                # Sales report sheet
                self._create_sales_sheet(writer, sales_report)
                
                # Profit report sheet
                self._create_profit_sheet(writer, profit_report)
                
                # Inventory report sheet
                self._create_inventory_sheet(writer, inventory_report)
                
                # Top products sheet
                self._create_top_products_sheet(writer, top_products)
            
            # Format the workbook
            self._format_workbook(filepath)
            
            return str(filepath)
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير Excel: {str(e)}")
    
    def _create_sales_sheet(self, writer, sales_report):
        """Create sales report sheet"""
        # Summary data
        summary_data = {
            'البيان': ['إجمالي المبيعات', 'عدد المعاملات', 'متوسط المعاملة'],
            'القيمة': [
                f"{sales_report['total_sales']:.2f} جنيه",
                sales_report['total_transactions'],
                f"{sales_report['average_transaction']:.2f} جنيه"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='تقرير المبيعات', index=False, startrow=0)
        
        # Sales details
        if sales_report['sales']:
            sales_data = []
            for sale in sales_report['sales']:
                sales_data.append({
                    'رقم الفاتورة': sale.invoice_no,
                    'التاريخ': sale.created_at.strftime('%Y-%m-%d'),
                    'الوقت': sale.created_at.strftime('%H:%M:%S'),
                    'العميل': sale.customer.name if sale.customer else 'عميل نقدي',
                    'المجموع الفرعي': sale.subtotal,
                    'الخصم': sale.discount,
                    'الضريبة': sale.tax,
                    'الإجمالي': sale.total,
                    'المدفوع': sale.paid,
                    'الباقي': sale.change,
                    'المستخدم': sale.user.name if sale.user else ''
                })
            
            sales_df = pd.DataFrame(sales_data)
            sales_df.to_excel(writer, sheet_name='تقرير المبيعات', index=False, startrow=6)
    
    def _create_profit_sheet(self, writer, profit_report):
        """Create profit report sheet"""
        profit_data = {
            'البيان': ['إجمالي الإيرادات', 'إجمالي التكلفة', 'صافي الربح', 'هامش الربح'],
            'القيمة': [
                f"{profit_report['total_revenue']:.2f} جنيه",
                f"{profit_report['total_cost']:.2f} جنيه", 
                f"{profit_report['profit']:.2f} جنيه",
                f"{profit_report['margin']:.2f}%"
            ]
        }
        
        profit_df = pd.DataFrame(profit_data)
        profit_df.to_excel(writer, sheet_name='تقرير الأرباح', index=False)
    
    def _create_inventory_sheet(self, writer, inventory_report):
        """Create inventory report sheet"""
        # Summary
        summary_data = {
            'البيان': ['إجمالي المنتجات', 'منتجات منخفضة', 'منتجات نفدت', 'قيمة المخزون'],
            'القيمة': [
                inventory_report['total_products'],
                inventory_report['low_stock'],
                inventory_report['out_of_stock'],
                f"{inventory_report['total_value']:.2f} جنيه"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='تقرير المخزون', index=False, startrow=0)
        
        # Products details
        if inventory_report['products']:
            products_data = []
            for product in inventory_report['products']:
                status = "نفد" if product.quantity <= 0 else "منخفض" if product.quantity <= product.min_quantity else "متوفر"
                value = product.cost_price * product.quantity
                
                products_data.append({
                    'الكود': product.sku,
                    'اسم المنتج': product.name_ar,
                    'الفئة': product.category.name_ar if product.category else '',
                    'الكمية': product.quantity,
                    'الحد الأدنى': product.min_quantity,
                    'سعر الشراء': product.cost_price,
                    'سعر البيع': product.sale_price,
                    'قيمة المخزون': value,
                    'الحالة': status,
                    'المزود': product.supplier.name if product.supplier else ''
                })
            
            products_df = pd.DataFrame(products_data)
            products_df.to_excel(writer, sheet_name='تقرير المخزون', index=False, startrow=6)
    
    def _create_top_products_sheet(self, writer, top_products):
        """Create top products sheet"""
        if top_products:
            products_data = []
            for i, (name, quantity, revenue) in enumerate(top_products, 1):
                products_data.append({
                    'الترتيب': i,
                    'اسم المنتج': name,
                    'الكمية المباعة': int(quantity),
                    'إجمالي الإيرادات': f"{float(revenue):.2f} جنيه"
                })
            
            products_df = pd.DataFrame(products_data)
            products_df.to_excel(writer, sheet_name='أكثر المنتجات مبيعاً', index=False)
    
    def _format_workbook(self, filepath):
        """Format Excel workbook for better presentation"""
        try:
            workbook = openpyxl.load_workbook(filepath)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format headers (first row)
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                
                # Add borders to all cells with data
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell.border = border
            
            workbook.save(filepath)
            
        except Exception as e:
            print(f"Warning: Could not format Excel file: {e}")
    
    def export_products(self, products, filename=None):
        """Export products to Excel"""
        try:
            if not filename:
                filename = f"منتجات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = self.output_dir / filename
            
            # Prepare data
            products_data = []
            for product in products:
                products_data.append({
                    'الكود': product.sku,
                    'اسم المنتج': product.name_ar,
                    'الوصف': product.description_ar or '',
                    'الفئة': product.category.name_ar if product.category else '',
                    'سعر الشراء': product.cost_price,
                    'سعر البيع': product.sale_price,
                    'الكمية': product.quantity,
                    'الحد الأدنى': product.min_quantity,
                    'الباركود': product.barcode or '',
                    'المزود': product.supplier.name if product.supplier else '',
                    'تاريخ الإضافة': product.created_at.strftime('%Y-%m-%d') if product.created_at else ''
                })
            
            df = pd.DataFrame(products_data)
            df.to_excel(str(filepath), index=False, sheet_name='المنتجات')
            
            # Format the file
            self._format_workbook(filepath)
            
            return str(filepath)
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير المنتجات: {str(e)}")
    
    def export_sales(self, sales, filename=None):
        """Export sales to Excel"""
        try:
            if not filename:
                filename = f"مبيعات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = self.output_dir / filename
            
            # Prepare data
            sales_data = []
            for sale in sales:
                sales_data.append({
                    'رقم الفاتورة': sale.invoice_no,
                    'التاريخ': sale.created_at.strftime('%Y-%m-%d'),
                    'الوقت': sale.created_at.strftime('%H:%M:%S'),
                    'العميل': sale.customer.name if sale.customer else 'عميل نقدي',
                    'رقم هاتف العميل': sale.customer.phone if sale.customer and sale.customer.phone else '',
                    'المجموع الفرعي': sale.subtotal,
                    'الخصم': sale.discount,
                    'الضريبة': sale.tax,
                    'الإجمالي': sale.total,
                    'المدفوع': sale.paid,
                    'الباقي': sale.change,
                    'المستخدم': sale.user.name if sale.user else '',
                    'الحالة': sale.status,
                    'ملاحظات': sale.notes or ''
                })
            
            df = pd.DataFrame(sales_data)
            df.to_excel(str(filepath), index=False, sheet_name='المبيعات')
            
            # Format the file
            self._format_workbook(filepath)
            
            return str(filepath)
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير المبيعات: {str(e)}")
    
    def export_repairs(self, repairs, filename=None):
        """Export repairs to Excel"""
        try:
            if not filename:
                filename = f"صيانة_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            filepath = self.output_dir / filename
            
            # Prepare data
            repairs_data = []
            for repair in repairs:
                repairs_data.append({
                    'رقم التذكرة': repair.ticket_no,
                    'العميل': repair.customer.name if repair.customer else '',
                    'رقم الهاتف': repair.customer.phone if repair.customer and repair.customer.phone else '',
                    'نوع الجهاز': repair.device_model,
                    'وصف المشكلة': repair.problem_desc,
                    'الحالة': repair.status,
                    'تاريخ الدخول': repair.entry_date.strftime('%Y-%m-%d %H:%M') if repair.entry_date else '',
                    'تاريخ الخروج': repair.exit_date.strftime('%Y-%m-%d %H:%M') if repair.exit_date else '',
                    'تكلفة القطع': repair.parts_cost,
                    'تكلفة العمالة': repair.labor_cost,
                    'التكلفة الإجمالية': repair.total_cost,
                    'المستخدم': repair.user.name if repair.user else '',
                    'ملاحظات': repair.notes or ''
                })
            
            df = pd.DataFrame(repairs_data)
            df.to_excel(str(filepath), index=False, sheet_name='الصيانة')
            
            # Format the file
            self._format_workbook(filepath)
            
            return str(filepath)
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير بيانات الصيانة: {str(e)}")
